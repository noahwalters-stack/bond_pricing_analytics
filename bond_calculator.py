"""
Bond Pricing Calculator
Author: Noah Walters

This project calculates bond prices, YTM, duration, and tax-equivalent yield
and exports those results directly into Excel."""

import argparse, math, pandas
from dataclasses import dataclass
from datetime import date
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

@dataclass
class Bond:
    bond_id: str
    bond_type: str
    face_value: float
    coupon_rate: float
    years_to_maturity: float
    frequency: int
    market_clean_price: float
    tax_rate: Optional[float]
    maturity_date: Optional[date]
    settlement_date: Optional[date]
    day_count: str

def price_from_ytm(face_value: float, coupon_rate: float, ytm: float, years_to_maturity: float, frequency: int)-> float:
    periods = max(1, int(round(years_to_maturity * frequency)))
    r = ytm / frequency
    coupon = face_value * coupon_rate / frequency
    price = 0.0
    for t in range(1, periods + 1):
        price += coupon / ((1 + r) ** t)
    price += face_value / ((1 + r) ** periods)
    return price

def get_ytm_from_price(price: float, face_value: float, coupon_rate: float, years_to_maturity: float, frequency: int)-> float:
    if price <= 0:
        raise ValueError('Price must be positive to solve the YTM.')
    
    def f(y: float)-> float:
        return price_from_ytm(face_value, coupon_rate, y, years_to_maturity, frequency) - price
    
    low, high = -0.99, 1.0
    f_low, f_high = f(low), f(high)
    attempts = 0
    while f_high > 0 and attempts < 20:
        high *= 2
        f_high = f(high)
        attempts += 1
    if f_low < 0:
        raise ValueError('Could not bracket YTM: price is too high for negative yields.')
    if f_high > 0:
        raise ValueError('Could not bracket YTM: price is too low for high yields.')
    for _ in range(200):
        mid = (low + high) / 2
        f_mid = f(mid)
        if abs(f_mid) < 1e-10:
            return mid
        if f_mid > 0:
            low = mid
        else:
            high = mid
    return (low + high) / 2

def duration(face_value: float, coupon_rate: float, ytm: float, years_to_maturity: float, frequency: int)-> Tuple[float, float]:
    periods = max(1, int(round(years_to_maturity * frequency)))
    r = ytm / frequency
    coupon = face_value * coupon_rate / frequency
    price = 0.0
    weighted_time = 0.0
    for t in range(1, periods + 1):
        cash_flow = coupon + (face_value if t == periods else 0.0)
        pv = cash_flow / ((1 + r) ** t)
        price += pv
        weighted_time += (t / frequency) * pv
    macaulay = weighted_time / price if price else 0.0
    modified = macaulay / (1 + r) if (1 + r) else 0.0
    return macaulay, modified

def normalize_columns(df: pandas.DataFrame)-> pandas.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip().lower().replace(" ", "_") for col in df.columns]
    rename_map = {
        "id": "bond_id",
        "bond": "bond_id",
        "type": "bond_type",
        "par": "face",
        "principal": "face",
        "coupon": "coupon_rate",
        "clean_price": "market_clean_price",
        "price": "market_clean_price",
        "maturity_years": "years_to_maturity",
        "years": "years_to_maturity",
        "freq": "frequency",
        "maturity": "maturity_date",
        "maturity_dt": "maturity_date",
        "settlement": "settlement_date",
        "settlement_dt": "settlement_date",
        "daycount": "day_count",
        "day_count_convention": "day_count",
    }
    return df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

def to_float(value: Any)-> Optional[float]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return float(value)


def to_date(value: Any)-> Optional[date]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    parsed = pandas.to_datetime(value, errors= "coerce")
    if pandas.isna(parsed):
        return None
    return parsed.date()

def normalize_day_count(value: Optional[str])-> str:
    if not value:
        return "30/360"
    cleaned = str(value).strip().upper().replace(" ", "")
    aliases = {
        "30/360US": "30/360",
        "30/360": "30/360",
        "ACT/ACT": "ACT/ACT",
        "ACTUAL/ACTUAL": "ACT/ACT",
        "ACT/360": "ACT/360",
        "ACTUAL/360": "ACT/360",
        "ACT/365": "ACT/365",
        "ACTUAL/365": "ACT/365",
        "ACT/365F": "ACT/365",
    }
    return aliases.get(cleaned, cleaned)

def add_months(base: date, months: int)-> date:
    return (pandas.Timestamp(base) + pandas.DateOffset(months=months)).date()

def day_count_30_360(start: date, end: date)-> int:
    d1 = 30 if start.day == 31 else start.day
    d2 = 30 if (end.day == 31 and d1 == 30) else end.day
    return (end.year - start.year) * 360 + (end.month - start.month) * 30 + (d2 - d1)

def year_fraction(start: date, end: date, convention: str)-> float:
    if end <= start:
        return 0.0
    if convention == "30/360":
        return day_count_30_360(start, end) / 360.0
    if convention == "ACT/360":
        return (end - start).days / 360.0
    if convention == "ACT/365":
        return (end - start).days / 365.0
    if convention == "ACT/ACT":
        total = 0.0
        cursor = start
        while cursor < end:
            next_year_start = date(cursor.year + 1, 1, 1)
            segment_end = min(next_year_start, end)
            days_in_year = 366 if pandas.Timestamp(cursor.year, 12, 31).is_leap_year else 365
            total += (segment_end - cursor).days / days_in_year
            cursor = segment_end
        return total
    return (end - start).days / 365.0

def previous_coupon_date(maturity: date, settlement: date, frequency: int)-> date:
    months = int(round(12 / frequency))
    current = maturity
    while current > settlement:
        current = add_months(current, -months)
    return current


def accrued_interest(face_value: float, coupon_rate: float, frequency: int, settlement: date, maturity: date, convention: str)-> float:
    prev_coupon = previous_coupon_date(maturity, settlement, frequency)
    next_coupon = add_months(prev_coupon, int(round(12 / frequency)))
    accrual = year_fraction(prev_coupon, settlement, convention)
    period = year_fraction(prev_coupon, next_coupon, convention)
    if period <= 0:
        return 0.0
    coupon = face_value * coupon_rate / frequency
    return coupon * (accrual / period)

def parse_bonds(df: pandas.DataFrame, default_tax_rate: Optional[float])-> List[Bond]:
    bonds: List[Bond] = []
    errors: List[str] = []
    for idx, row in df.iterrows():
        bond_id = str(row.get("bond_id") or row.get("id") or f"Bond-{idx + 1}")
        bond_type = str(row.get("bond_type") or "corporate").strip().lower()
        if bond_type in {"corp", "corporate"}:
            bond_type = "corporate"
        elif bond_type in {"muni", "municipal", "municipality"}:
            bond_type = "municipal"
        try:
            face_value = to_float(row.get("face")) or 100.0
            coupon_rate = to_float(row.get("coupon_rate"))
            years_to_maturity = to_float(row.get("years_to_maturity"))
            frequency = int(to_float(row.get("frequency")) or 2)
            market_clean_price = to_float(row.get("market_clean_price"))
            tax_rate = to_float(row.get("tax_rate"))
            maturity_date = to_date(row.get("maturity_date"))
            settlement_date = to_date(row.get("settlement_date"))
            day_count = normalize_day_count(row.get("day_count"))
            if coupon_rate is None:
                raise ValueError("coupon_rate is required")
            if years_to_maturity is None:
                if maturity_date is None:
                    raise ValueError("years_to_maturity or maturity_date is required")
                if settlement_date is None:
                    raise ValueError("settlement_date is required when using maturity_date")
                if settlement_date >= maturity_date:
                    raise ValueError("settlement_date must be before maturity_date")
                years_to_maturity = year_fraction(settlement_date, maturity_date, day_count)
            if market_clean_price is None:
                raise ValueError("market_clean_price is required")
            if frequency <= 0:
                raise ValueError("frequency must be positive")
            if tax_rate is None:
                tax_rate = default_tax_rate
            bonds.append(
                Bond(
                    bond_id=bond_id,
                    bond_type=bond_type,
                    face_value=face_value,
                    coupon_rate=coupon_rate,
                    years_to_maturity=years_to_maturity,
                    frequency=frequency,
                    market_clean_price=market_clean_price,
                    tax_rate=tax_rate,
                    maturity_date=maturity_date,
                    settlement_date=settlement_date,
                    day_count=day_count,
                )
            )
        except Exception as exc:
            errors.append(f"{bond_id}: {exc}")
    if errors:
        raise ValueError("Invalid bond rows:\n" + "\n".join(errors))
    return bonds

def compute_bond_metrics(bond: Bond)-> Dict[str, Any]:
    accrued = 0.0
    if bond.maturity_date and bond.settlement_date:
        accrued = accrued_interest(
            bond.face_value,
            bond.coupon_rate,
            bond.frequency,
            bond.settlement_date,
            bond.maturity_date,
            bond.day_count,
        )
    price_dirty = bond.market_clean_price + accrued
    ytm_from_price = get_ytm_from_price(
        price_dirty,
        bond.face_value,
        bond.coupon_rate,
        bond.years_to_maturity,
        bond.frequency,
    )
    ytm_quoted = round(ytm_from_price, 7)
    price_from_yield = price_from_ytm(
        bond.face_value,
        bond.coupon_rate,
        ytm_quoted,
        bond.years_to_maturity,
        bond.frequency,
    )
    ytm = ytm_from_price
    if price_dirty is None or ytm is None:
        raise ValueError(f"Unable to compute price and YTM for bond {bond.bond_id}.")
    macaulay, modified = duration(
        bond.face_value,
        bond.coupon_rate,
        ytm,
        bond.years_to_maturity,
        bond.frequency,
    )
    price_clean = price_dirty - accrued
    coupon_payment = bond.face_value * bond.coupon_rate / bond.frequency
    tax_equivalent_yield = None
    if bond.bond_type == "municipal" and bond.tax_rate is not None:
        if bond.tax_rate >= 1:
            raise ValueError(f"Invalid tax_rate for bond {bond.bond_id}.")
        tax_equivalent_yield = ytm / (1 - bond.tax_rate)
    return {
        "bond_id": bond.bond_id,
        "bond_type": bond.bond_type,
        "face": bond.face_value,
        "coupon_rate": bond.coupon_rate,
        "years_to_maturity": bond.years_to_maturity,
        "frequency": bond.frequency,
        "clean_price": price_clean,
        "coupon_payment": coupon_payment,
        "dirty_price": price_dirty,
        "accrued_interest": accrued,
        "ytm": ytm,
        "macaulay_duration": macaulay,
        "modified_duration": modified,
        "tax_rate": bond.tax_rate,
        "tax_equivalent_yield": tax_equivalent_yield,
        "price_from_ytm": price_from_yield,
        "ytm_from_price": ytm_from_price,
        "price_diff": None if price_from_yield is None else price_dirty - price_from_yield,
        "maturity_date": bond.maturity_date,
        "settlement_date": bond.settlement_date,
        "day_count": bond.day_count,
    }

def build_summary(df: pandas.DataFrame)-> pandas.DataFrame:
    metrics = [
        ("Total Bonds", len(df)),
        ("Average Dirty Price", df["dirty_price"].mean()),
        ("Average YTM", df["ytm"].mean()),
        ("Average Macaulay Duration", df["macaulay_duration"].mean()),
        ("Average Modified Duration", df["modified_duration"].mean()),
    ]
    return pandas.DataFrame(metrics, columns=["Metric", "Value"])

def sensitivity_blocks(df: pandas.DataFrame, bps_range: int, bps_step: int)-> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    step = bps_step / 10000
    span = bps_range / 10000
    for _, row in df.iterrows():
        base = row["ytm"]
        yields = [base + delta for delta in frange(-span, span, step)]
        yields = [max(-0.95, y) for y in yields]
        prices = [
            price_from_ytm(
                row["face"],
                row["coupon_rate"],
                y,
                row["years_to_maturity"],
                int(row["frequency"]),
            )
            for y in yields
        ]
        blocks.append({"bond_id": row["bond_id"], "yields": yields, "prices": prices})
    return blocks

def frange(start: float, stop: float, step: float)-> Iterable[float]:
    current = start
    while current <= stop + 1e-12:
        yield current
        current += step


def apply_number_formats(ws, start_row: int, end_row: int)-> None:
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=1).number_format = "0.00%"
        ws.cell(row=row, column=2).number_format = "$#,##0.00"


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            value_length = 10 if isinstance(cell.value, date) else len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        if max_length:
            ws.column_dimensions[column_letter].width = max_length + 2

def apply_bonds_number_formats(ws)-> None:
    header_map = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}
    formats = {
        "years_to_maturity": "0.000",
        "coupon_payment": "0.00",
        "clean_price": "0.000",
        "accrued_interest": "0.000",
        "dirty_price": "0.000",
        "ytm": "0.0000%",
        "tax_equivalent_yield": "0.0000%",
        "macaulay_duration": "0.000",
        "modified_duration": "0.000",
        "price_from_ytm": "0.000",
        "ytm_from_price": "0.0000%",
        "price_diff": "0.0000",
    }
    for header, number_format in formats.items():
        col_idx = header_map.get(header)
        if not col_idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = number_format

def write_excel_report(bonds_df: pandas.DataFrame, summary_df: pandas.DataFrame, blocks: List[Dict[str, Any]], output_path: str)-> None:
    with pandas.ExcelWriter(output_path, engine="openpyxl") as writer:
        bonds_df.to_excel(writer, sheet_name="Bonds", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        pandas.DataFrame().to_excel(writer, sheet_name="Sensitivity", index=False)
        pandas.DataFrame().to_excel(writer, sheet_name="Charts", index=False)
    wb = load_workbook(output_path)
    ws_bonds = wb["Bonds"]
    ws_sensitivity = wb["Sensitivity"]
    ws_charts = wb["Charts"]
    apply_bonds_number_formats(ws_bonds)
    header_map = {ws_bonds.cell(row=1, column=col).value: col for col in range(1, ws_bonds.max_column + 1)}
    for date_col in ("maturity_date", "settlement_date"):
        col_idx = header_map.get(date_col)
        if col_idx:
            ws_bonds.column_dimensions[get_column_letter(col_idx)].width = 14
            for row in range(2, ws_bonds.max_row + 1):
                cell = ws_bonds.cell(row=row, column=col_idx)
                if isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"
    chart_row = 1
    for block in blocks:
        start_row = ws_sensitivity.max_row + 2 if ws_sensitivity.max_row > 1 else 1
        ws_sensitivity.cell(row=start_row, column=1, value=f"Bond {block['bond_id']} Price Sensitivity").font = Font(bold=True)
        header_row = start_row + 1
        ws_sensitivity.cell(row=header_row, column=1, value="Yield").font = Font(bold=True)
        ws_sensitivity.cell(row=header_row, column=2, value="Price").font = Font(bold=True)
        data_start = header_row + 1
        for i, (yield_value, price_value) in enumerate(zip(block["yields"], block["prices"])):
            ws_sensitivity.cell(row=data_start + i, column=1, value=yield_value)
            ws_sensitivity.cell(row=data_start + i, column=2, value=price_value)
        data_end = data_start + len(block["yields"]) - 1
        apply_number_formats(ws_sensitivity, data_start, data_end)
        chart = LineChart()
        chart.title = f"{block['bond_id']} Price vs Yield"
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Yield"
        data_ref = Reference(ws_sensitivity, min_col=2, min_row=header_row, max_row=data_end)
        cats_ref = Reference(ws_sensitivity, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 7
        chart.width = 15
        ws_charts.add_chart(chart, f"A{chart_row}")
        chart_row += 16
    autosize_worksheet(ws_bonds)
    autosize_worksheet(ws_sensitivity)
    autosize_worksheet(ws_charts)
    autosize_worksheet(wb["Summary"])
    wb.save(output_path)

def load_input(path: str)-> pandas.DataFrame:
    return pandas.read_excel(path) if path.lower().endswith((".xlsx", ".xls")) else pandas.read_csv(path)

def main()-> None:
    parser = argparse.ArgumentParser(description="Bond price calculator with YTM and duration reporting.")
    parser.add_argument("--input", required=True, help="Path to CSV or Excel bond input file.")
    parser.add_argument("--output", default="bond_report.xlsx", help="Output Excel report path.")
    parser.add_argument("--tax-rate", type=float, default=None, help="Default tax rate for municipal bonds.")
    parser.add_argument("--sensitivity-bps", type=int, default=200, help="Yield shock range in bps.")
    parser.add_argument("--sensitivity-step-bps", type=int, default=25, help="Yield step size in bps.")
    args = parser.parse_args()
    df = normalize_columns(load_input(args.input))
    bonds = parse_bonds(df, args.tax_rate)
    metrics = [compute_bond_metrics(bond) for bond in bonds]
    bonds_df_full = pandas.DataFrame(metrics)
    output_cols = [
        "bond_id",
        "bond_type",
        "years_to_maturity",
        "coupon_payment",
        "clean_price",
        "accrued_interest",
        "dirty_price",
        "ytm",
        "tax_equivalent_yield",
        "macaulay_duration",
        "modified_duration",
        "price_from_ytm",
        "ytm_from_price",
        "price_diff",
    ]
    bonds_df_output = bonds_df_full[output_cols]
    summary_df = build_summary(bonds_df_full)
    blocks = sensitivity_blocks(bonds_df_full, args.sensitivity_bps, args.sensitivity_step_bps)
    write_excel_report(bonds_df_output, summary_df, blocks, args.output)
    print(f"Report written to {args.output}")

if __name__ == "__main__":
    main()
